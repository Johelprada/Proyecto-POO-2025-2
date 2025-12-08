"""
                                                                        COSAS A PRACTICAR
La parte del codigo que trabaja las especificaciones del excel esta casi que calcada de tutoriales y 
post de reddit, debo investigar mas sobre ella para poder justificarla correctamente
"""
"""
                                                                        COSAS A AGREGAR
        Tener en cuenta pep8
        Añadir encapsulamiento
        seguridad
        especializarlo  tal vez
        Intentar para tener colas
        Q hago si me meten 4k datos y no puedo procesarlos o si el historial se vuelve muy grande?
"""
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# Clase que representa un producto individual en el inventario
class Producto:
    """
    Cada producto tiene: codigo, nombre, precio, stock, categoria y fecha de creacion que es lo que estamos asignando aquí.
    """
    # Metodo constructor que se ejecuta al crear un nuevo objeto Producto
    def __init__(self, codigo: str, nombre: str, precio: float, stock: int, categoria: str):
        # Asignamos el codigo, nombre, precio, categoria, stock, y fecha de creacion del producto 
        self.codigo = codigo
        self.nombre = nombre
        self.precio = precio
        self.stock = stock
        self.categoria = categoria
        self.fecha_creacion = datetime.now()
         #self.proveedor = ""  

    # Metodo para modificar el stock del producto
    def modificar_stock(self, cantidad: int, operacion: str):
        stock_anterior = self.stock  # Guardamos el stock anterior para el historial
        
        if operacion == 'agregar':
            self.stock += cantidad
            accion = "agrego"
        elif operacion == 'retirar':
            # Verificamos que haya suficiente stock disponible
            if cantidad > self.stock:
                print(f"Error: Solo hay {self.stock} unidades disponibles")
                return False
             # Restamos la cantidad del stock actual
            self.stock -= cantidad
            accion = "retiro"
            #Si la operacion no es valida
        else:
            print("Operacion no valida")
            return False
        
        # Retorna el mensaje de cambio para que el historial lo registre
        cambio = f"Se {accion} {cantidad} unidades de '{self.nombre}' (Stock: {stock_anterior} -> {self.stock})"
        print(f"Operacion exitosa: {cambio}")
        return True, cambio  
    
    # Metodo para actualizar el precio del producto
    def actualizar_precio(self, nuevo_precio: float):
        #guardamos el precio actual antes de cambiarlo al nuevo
        precio_anterior = self.precio
        self.precio = nuevo_precio
        
       # Retorna el mensaje de cambio para que el historial lo registre
        cambio = f"Precio de '{self.nombre}' actualizado: ${precio_anterior:.2f} -> ${nuevo_precio:.2f}"
        print(f"Precio actualizado: {cambio}")
        return cambio  
         #def calcular_valor_total(self):
    #    return self.precio * self.stock
    
    # Definimos como se muestra el producto cuando usamos print
    def __str__(self):
        return f"[{self.codigo}] {self.nombre} - Stock: {self.stock} - Precio: ${self.precio:.2f} - Categoria: {self.categoria}"


# Clase que gestiona el catalogo completo de productos
class Catalogo:
    """
     Administra todos los productos del inventario y encapsula el diccionario de productos y sus operaciones
    """
    def __init__(self):
        #  Diccionario privado que almacena todos los productos
        self._productos = {}
    
    # Funcion para crear un nuevo producto y agregarlo al catalogo
    def agregar_producto(self, codigo: str, nombre: str, precio: float, stock: int, categoria: str):
        # Verificamos si ya existe un producto con ese codigo en el catalogo
        if codigo in self._productos:
            print(f"Error: Ya existe un producto con el codigo '{codigo}'")
            return None
        
        # Verificamos que el precio y el stock no sean numeros negativos
        if precio < 0 or stock < 0:
            print("Error: El precio y el stock no pueden ser negativos")
            return None
        
        # Creamos un nuevo objeto Producto con los datos proporcionados
        producto = Producto(codigo, nombre, precio, stock, categoria)
        
        # Agregamos el producto al diccionario usando el codigo como clave
        self._productos[codigo] = producto
        
        print(f"Producto '{nombre}' creado exitosamente")
        
        #Retorna el producto y un mensaje para el historial
        return producto, f"Producto creado: {nombre} (Codigo: {codigo})"
    
    def buscar_producto(self, codigo: str):
         # Usamos el metodo get del diccionario que retorna el valor 
        return self._productos.get(codigo, None)
    
    # Metodo para eliminar un producto del catalogo
    def eliminar_producto(self, codigo: str):
        """
        Elimina un producto del catalogo usando su codigo y retorna el mensaje de eliminacion 
        """
        # Verifica si el producto existe en el catalogo
        if codigo not in self._productos:
            print(f"Error: No existe un producto con el codigo '{codigo}'")
            return None
        
        # Obtenemos el producto antes de eliminarlo
        producto = self._productos[codigo]
        nombre_producto = producto.nombre
        
        #  Elimina el producto del diccionario
        del self._productos[codigo]
        
        print(f"Producto '{nombre_producto}' eliminado exitosamente")
        
        # Esto retorna el mensaje para el historial
        return f"Producto eliminado: {nombre_producto} (Codigo: {codigo})"
    
    def listar_productos(self):

        # Verificamos si el diccionario de productos esta vacio
        if not self._productos:
            print("El catalogo esta vacio")
            return
        
        print("\n" + "="*70)
        print("CATALOGO DE PRODUCTOS")
        print("="*70)
        
          # Recorremos todos los productos del diccionario e imprimimos cada profucto
        for producto in self._productos.values():
            print(producto)
        print("="*70)
    
    #  Metodo para obtener todos los productos al exportarlo a excel

    def obtener_todos(self):
     
        return list(self._productos.values())


# Clase que gestiona las categorias del sistema
class GestorCategorias:
    """
     administra todas las categorias disponibles y permite crear, listar y validar categorias
    """
    def __init__(self):
        # Lista privada que almacena las categorias
        self._categorias = []
    
    # Funcion para crear categorias y modificarlas
    def crear_categoria(self, nombre_categoria: str = None):
        """
        Permite al usuario crear una nueva categoria y la agrega a la lista de categorias existentes
        """
        #  Algunas corroboraciones, como que se le de nombre, este vacia o si ya existe
        if nombre_categoria is None:
            print("\n--- CREAR NUEVA CATEGORIA ---")
            nombre_categoria = input("Nombre de la categoria: ").strip()
        
        if not nombre_categoria:
            print("Error: El nombre de la categoria no puede estar vacio")
            return None
        
        if nombre_categoria.lower() in [cat.lower() for cat in self._categorias]:
            print(f"Error: La categoria '{nombre_categoria}' ya existe")
            return None
        
        # Agregamos la categoria a la lista
        self._categorias.append(nombre_categoria)
        
        print(f"Categoria '{nombre_categoria}' creada exitosamente")
        
        #  Retornamos la categoria y el mensaje para el historial
        return nombre_categoria, f"Categoria creada: {nombre_categoria}"
    
    def listar_categorias(self):
        """
        Muestra todas las categorias disponibles en el sistema
        """
        if not self._categorias:
            print("\nNo hay categorias registradas aun")
            return
        
        print("\n" + "="*50)
        print("CATEGORIAS DISPONIBLES")
        print("="*50)
        for i, categoria in enumerate(self._categorias, 1):
            print(f"{i}. {categoria}")
        print("="*50)
    
    def seleccionar_categoria(self):
        """
        FUncion para crear una categoria nueva si no hay categorias, obliga a crear una
        """
        if not self._categorias:
            print("\nNo hay categorias disponibles. Debes crear una primero.")
            return self.crear_categoria()
        
        # Muestra las categorias existentes
        print("\n--- SELECCIONAR CATEGORIA ---")
        print("Categorias disponibles:")
        for i, categoria in enumerate(self._categorias, 1):
            print(f"{i}. {categoria}")
        print(f"{len(self._categorias) + 1}. Crear nueva categoria")
        
        opcion = input("\nSelecciona una opcion: ").strip()
        
        # Verifica si eligio crear nueva categoria o una ya existente
        if opcion == str(len(self._categorias) + 1):
            return self.crear_categoria()
        
        try:
            indice = int(opcion) - 1
            if 0 <= indice < len(self._categorias):
                #  Retornamos la categoria seleccionada
                return self._categorias[indice], None
            else:
                print("Opcion no valida")
                return None, None
        except ValueError:
            print("Opcion no valida")
            return None, None
    
    # metodo para verificar si hay categorias
    def tiene_categorias(self):

        """Retorna True si hay al menos una categoria"""
        return len(self._categorias) > 0


#  Clase que gestiona el historial de cambios del sistema
class Historial:
    """
    Aqui se registran los cambios realizados con fecha y hora de estos
    """
    def __init__(self):
        # Lista privada que almacena los cambios
        self._cambios = []
    
    #  Metodo para registrar un nuevo cambio
    def registrar_cambio(self, descripcion: str):

        """se agrega un cambio al historial con fecha y hora"""

        cambio_completo = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {descripcion}"
        self._cambios.append(cambio_completo)
    
    def ver_historial(self):
        # Verificamos si la lista de cambios esta vacia
        if not self._cambios:
            print("No hay cambios registrados aun")
            return
        
        print("\n" + "="*70)
        print("HISTORIAL DE CAMBIOS")
        print("="*70)
        
        # Recorremos la lista de cambios con enumerate para obtener el indice
        for i, cambio in enumerate(self._cambios, 1):
            print(f"{i}. {cambio}")
        
        print("="*70)
    
    #  Metodo para obtener todos los cambios (no lo estoy usando del todo aun)
    def obtener_cambios(self):
        """ Retorna una copia de la lista de cambios"""
        return self._cambios.copy()


# Clase que maneja la exportacion de datos a Excel
class ExportadorExcel:
    """
    Se encarga de realizar el manejo de los csv y creacion del archivo excel     
    """
    
    def exportar_catalogo(self, productos: list):
         # Verificamos si hay productos en el catalogo
        if not productos:
            print("No hay productos para exportar")
            return None
        
        # Definimos el nombre del archivo CSV temporal
        archivo_csv = 'catalogo_temp.csv'
        datos = [['Codigo', 'Nombre', 'Precio', 'Stock', 'Categoria', 'Fecha_Creacion']]
        
        # Recorremos todos los productos del catalogo y creamos una lista con estos datos
        for producto in productos:
            fila = [
                producto.codigo,                                       
                producto.nombre,                                        
                producto.precio,                                        
                producto.stock,                                        
                producto.categoria,                                     
                producto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')  
            ]
             # Agregamos la fila a la lista de datos
            datos.append(fila)

        # Abrimos el archivo CSV
        with open(archivo_csv, 'w', newline='', encoding='utf-8') as f:
            escritor = csv.writer(f)
            escritor.writerows(datos)

        # Creamos un nuevo libro de excel, obtenemos una hoja activa y le asignamos el nombre/ 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Abrimos el archivo CSV que creamos anteriormente y con el lector de csv convertimos todo el contenido a una lista
        with open(archivo_csv, 'r', encoding='utf-8') as f:
            lector = csv.reader(f)
            datos_csv = list(lector)
        
        # aquí no le entendi muy bien al tutorial sobre como funciona, pero funciona
        for fila_num, fila_datos in enumerate(datos_csv, 1):
            for col_num, valor in enumerate(fila_datos, 1):
                if fila_num > 1 and col_num in [3, 4]:
                    valor = float(valor) if col_num == 3 else int(valor)
                ws.cell(row=fila_num, column=col_num, value=valor)
        
        # Aplica el formato a los encabezados
        self._aplicar_formato(ws)
        
        # Definimos el nombre del archivo Excel final y lo guardamos en la carpeta donde esta el codigo
        archivo_excel = 'catalogo_productos.xlsx'
        wb.save(archivo_excel)
        print(f"Catalogo exportado exitosamente a '{archivo_excel}'")
        
        return "Se exporto el catalogo a Excel"
    
    #  Metodo para aplicar formato al Excel
    def _aplicar_formato(self, ws):
        """ Aplica formato de colores y alineacion a los encabezados"""
         # Modifica el formato con el cual se va a ver el excel
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        alineacion = Alignment(horizontal="center", vertical="center")
        
         # Recorre las primeras 6 columnas para aplicar los cambios
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = font_encabezado
            cell.fill = fill_encabezado
            cell.alignment = alineacion

        # Ajusta automaticamente el ancho de cada columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


# Clase principal que coordina todo el sistema de inventario
class SistemaInventario:
    """
    Gestiona las demas clases para hacer que el sistema funcione correctamente
    """
    def __init__(self):
    #Inicializamos todas las clases del sistema

        self.catalogo = Catalogo()
        self.categorias = GestorCategorias()
        self.historial = Historial()
        self.exportador = ExportadorExcel()
    
    def mostrar_menu(self):
        print("\n" + "="*50)
        print("SISTEMA DE INVENTARIO v2.2")
        print("="*50)
        print("1. Crear nuevo producto")
        print("2. Ver catalogo completo")
        print("3. Modificar stock de producto")
        print("4. Actualizar precio de producto")
        print("5. Ver historial de cambios")
        print("6. Exportar catalogo a Excel")
        print("7. Gestionar categorias")
        print("8. Eliminar producto")
        print("9. Salir")
        print("="*50)
    
    def menu_crear_producto(self):
        print("\n--- CREAR NUEVO PRODUCTO ---")
        
         # Solicita los datos del producto y eliminamos espacios en blanco o converrtimos los numeros a numeros decimales o enteros segun sea necesairio
        codigo = input("Codigo del producto: ").strip()
        nombre = input("Nombre: ").strip()
        precio = float(input("Precio: $"))
        stock = int(input("Stock inicial: "))
        
        # Aqui se usa la funcion de seleccionar categoria y si no se selecciono o creo una categoria cancelamos la creacion
        resultado = self.categorias.seleccionar_categoria()
        
        #Verifica el resultado de seleccionar categoria
        if resultado is None:
            print("Creacion de producto cancelada")
            return
        
        #  Desempaqueta las categoria y mensajes del historial
        categoria, mensaje_cat = resultado
        
        if not categoria:
            print("Creacion de producto cancelada")
            return
        
        # Si se creo una categoria nueva, registramos en historial
        if mensaje_cat:
            self.historial.registrar_cambio(mensaje_cat)
        
        # Crea el producto y lo registra en historial
        resultado_producto = self.catalogo.agregar_producto(codigo, nombre, precio, stock, categoria)
        if resultado_producto:
            producto, mensaje = resultado_producto
            self.historial.registrar_cambio(mensaje)
    
    def menu_modificar_stock(self):
        print("\n--- MODIFICAR STOCK ---")
        
         # Buscamos el producto en el catalogo y verificamos si se encuentra
        codigo = input("Codigo del producto: ").strip()
        producto = self.catalogo.buscar_producto(codigo)
        
        if not producto:
            print(f"No se encontro producto con codigo '{codigo}'")
            return
        
        print(f"Producto encontrado: {producto}")
        
        print("\n1. Agregar stock")
        print("2. Retirar stock")
        opcion = input("Selecciona operacion: ").strip()
        
        # Solicitamos la cantidad, la convertimos a entero y verificamos algunas cosas
        cantidad = int(input("Cantidad: "))
        
        if cantidad <= 0:
            print("La cantidad debe ser mayor a 0")
            return
        
        #  Ejecuta la operacion y registra en el historial si fue exitosa
        if opcion == "1":
            resultado = producto.modificar_stock(cantidad, 'agregar')
            if resultado and resultado[0]:
                self.historial.registrar_cambio(resultado[1])
        elif opcion == "2":
            resultado = producto.modificar_stock(cantidad, 'retirar')
            if resultado and resultado[0]:
                self.historial.registrar_cambio(resultado[1])
        else:
            print("Opcion no valida")
    
    def menu_actualizar_precio(self):
        print("\n--- ACTUALIZAR PRECIO ---")
        
        codigo = input("Codigo del producto: ").strip()
        producto = self.catalogo.buscar_producto(codigo)
        
          # Verificamos si se encontro el producto
        if not producto:
            print(f"No se encontro producto con codigo '{codigo}'")
            return
        
        print(f"Producto encontrado: {producto}")
        
        # Solicitamos el nuevo precio y lo convertimos a decimal
        nuevo_precio = float(input("Nuevo precio: $"))
        
        if nuevo_precio < 0:
            print("El precio no puede ser negativo")
            return
        
        # Actualizamos precio y registramos en historial
        mensaje = producto.actualizar_precio(nuevo_precio)
        self.historial.registrar_cambio(mensaje)
    
    #  Menu para gestionar categorias
    def menu_gestionar_categorias(self):
       
        while True:
            print("\n" + "="*50)
            print("GESTIONAR CATEGORIAS")
            print("="*50)
            print("1. Ver categorias existentes")
            print("2. Crear nueva categoria")
            print("3. Volver al menu principal")
            print("="*50)
            
            opcion = input("\nSelecciona una opcion (1-3): ").strip()
            
            if opcion == "1":
                self.categorias.listar_categorias()
            elif opcion == "2":
                #  Crea una categoria y la registra en el historial
                resultado = self.categorias.crear_categoria()
                if resultado:
                    categoria, mensaje = resultado
                    self.historial.registrar_cambio(mensaje)
            elif opcion == "3":
                break
            else:
                print("Opcion no valida. Elige entre 1-3")
            
            input("\nPresiona Enter para continuar...")
    
    #  metodo para eliminar un producto
    def menu_eliminar_producto(self):
      
        print("\n--- ELIMINAR PRODUCTO ---")
        
        #  Solicita el codigo del producto a eliminar
        codigo = input("Codigo del producto a eliminar: ").strip()
        
        # Busca el producto en el catalogo y verifica que exista
        producto = self.catalogo.buscar_producto(codigo)
        
        if not producto:
            print(f"No se encontro producto con codigo '{codigo}'")
            return
        
        #  Mostramos la informacion del producto encontrado y solicitamos confirmacion
        print(f"\nProducto encontrado:")
        print(producto)
        
        confirmacion = input("\n¿Estas seguro de eliminar este producto? (s/n): ").strip().lower()
        
        # si confirma lo eliminarmos y registramos en el historial
        if confirmacion == 's':
            mensaje = self.catalogo.eliminar_producto(codigo)
            if mensaje:
                self.historial.registrar_cambio(mensaje)
        else:
            print("Eliminacion cancelada")
    
    #  Menu para exportar catalogo a Excel
    def menu_exportar(self):

        """Exporta el catalogo y registra en historial"""
        productos = self.catalogo.obtener_todos()
        mensaje = self.exportador.exportar_catalogo(productos)
        if mensaje:
            self.historial.registrar_cambio(mensaje)
    
    # Metodo principal que ejecuta el sistema
    def ejecutar(self):
        print("Bienvenido al Sistema de Inventario!")
        
        # Iniciamos un ciclo infinito para el menu
        while True:
            self.mostrar_menu()
            
            opcion = input("\nSelecciona una opcion (1-9): ").strip()
            
            if opcion == "1":
                self.menu_crear_producto()
            elif opcion == "2":
                self.catalogo.listar_productos()
            elif opcion == "3":
                self.menu_modificar_stock()
            elif opcion == "4":
                self.menu_actualizar_precio()
            elif opcion == "5":
                self.historial.ver_historial()
            elif opcion == "6":
                self.menu_exportar()
            elif opcion == "7":
                self.menu_gestionar_categorias()
            elif opcion == "8":
                self.menu_eliminar_producto()
            elif opcion == "9":
                print("\nGracias por usar el sistema de inventario!")
                break

            # Si elige cualquier otra opcion
            else:
                print("Opcion no valida. Elige entre 1-9")
            
            input("\nPresiona Enter para continuar...")


#  Punto de entrada del programa
if __name__ == "__main__":
    # Crea una instancia del sistema y lo ejecutamos
    sistema = SistemaInventario()
    sistema.ejecutar()