"""
                                                                        COSAS A PRACTICAR
La parte del codigo que trabaja las especificaciones del excel esta casi que calcada de tutoriales y 
post de reddit, debo investigar mas sobre ella para poder justificarla correctamente
"""
"""
                                                                        COSAS A AGREGAR
        Tener en cuenta pep8
        Añadir clases
        Añadir encapsulamiento
        sewguridad
        especializarlo  tal vez
        Que recuerde los categorias del registro o que tenga algunas predefinidas, tal vez que el usuario pueda crear unas categorias y luego a la hora de crear un producto
   se pueda usar, tipo, crear producto, poner las cosas y cuando llegue a categoria poder elegir entre: usar categoria ya existente o crear una nueva,
   si no existen categorias existentes obligar a crear una
        Intentar para tener colas
        funcion para eliminar productos
        Q hago si me meten 4k datos y no puedo procesarlos o si el historial se vuelve muy grande?
        Intentar que la funcion que crea, muestra y trabaja el historial y catalogo quede como una clase
    """
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

catalogo_productos = {}  #dicionario el cual va a almacenar todos los productos del catalogo
historial_cambios = [] # lista en la cual guardaremos el historial de todas las operaciones
#total_productos = 0  
#ultima_modificacion = None


# Definimos la clase Producto
class Producto:
    """
    Cada producto tiene: codigo, nombre, precio, stock, categoria y fecha de creacion que es lo que estamos asignando aquí.
    """
    # Metodo constructor que se ejecuta al crear un nuevo objeto Producto
    def __init__(self, codigo: str, nombre: str, precio: float, stock: int, categoria: str):

        # Asignamos el codigo, nombre, precio,categoria,stock,y fecha de creacion del producto 
        self.codigo = codigo
        self.nombre = nombre
        self.precio = precio
        self.stock = stock
        self.categoria = categoria
        self.fecha_creacion = datetime.now()
        #self.proveedor = ""  #No se si deberia seguir con esto
    
    # Metodo para modificar el stock del producto
    def modificar_stock(self, cantidad: int, operacion: str):

        stock_anterior = self.stock # Guardamos el stock anterior para el historial
        
        if operacion == 'agregar':
            self.stock += cantidad
            accion = "agrego"
        elif operacion == 'retirar':
            # Verificamos que haya suficiente stock disponible
            if cantidad > self.stock:
                print(f"Error: Solo hay {self.stock} unidades disponibles")
                return False
            # Restamos la cantidad del stock actual
            self.stock -= cantidad
            accion = "retiro"
        # Si la operacion no es valida
        else:
            print("Operacion no valida")
            return False
        
        # Creamos un mensaje del cambio realizado con fecha y hora y luego lo agregamos al historial
        cambio = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Se {accion} {cantidad} unidades de '{self.nombre}' (Stock: {stock_anterior} -> {self.stock})"
        historial_cambios.append(cambio)
        print(f"Operacion exitosa: {cambio}")
        return True
    
    # Metodo para actualizar el precio del producto
    def actualizar_precio(self, nuevo_precio: float):
        # Guardamos el precio actual antes de cambiarlo
        precio_anterior = self.precio
        # Asignamos el nuevo precio al producto
        self.precio = nuevo_precio
        
        # repetimos el mismo proceso hecho en el anterior metodo
        cambio = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Precio de '{self.nombre}' actualizado: ${precio_anterior:.2f} -> ${nuevo_precio:.2f}"
        historial_cambios.append(cambio)
        print(f"Precio actualizado: {cambio}")
    
    #def calcular_valor_total(self):
    #    return self.precio * self.stock
    
    # Definimos como se muestra el producto cuando usamos print
    def __str__(self):
        # mostramos la informacion del producto
        return f"[{self.codigo}] {self.nombre} - Stock: {self.stock} - Precio: ${self.precio:.2f} - Categoria: {self.categoria}"


# Funcion para crear un nuevo producto y agregarlo al catalogo
def crear_producto(codigo: str, nombre: str, precio: float, stock: int, categoria: str):
 
    # Verificamos si ya existe un producto con ese codigo en el catalogo
    if codigo in catalogo_productos:
        print(f"Error: Ya existe un producto con el codigo '{codigo}'")
        return None
    
    # Verificamos que el precio y el stock no sean numeros negativos
    if precio < 0 or stock < 0:
        print("Error: El precio y el stock no pueden ser negativos")
        return None
    
    # Creamos un nuevo objeto Producto con los datos proporcionados
    producto = Producto(codigo, nombre, precio, stock, categoria)
    
    # Agregamos el producto al diccionario usando el codigo como clave
    catalogo_productos[codigo] = producto
    
    # Creamos un mensaje con la fecha, hora y detalles de la creacion y agregamos el el cambio al historial
    cambio = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Producto creado: {nombre} (Codigo: {codigo})"
    historial_cambios.append(cambio)
    print(f"Producto '{nombre}' creado exitosamente")
    
    return producto


# Funcion para buscar un producto en el catalogo por su codigo
def buscar_producto(codigo: str):
    # Usamos el metodo get del diccionario que retorna el valor 
    return catalogo_productos.get(codigo, None)

# Funcion para mostrar todos los productos del catalogo
def listar_productos():

    # Verificamos si el diccionario de productos esta vacio
    if not catalogo_productos:
        print("El catalogo esta vacio")
        return
    
    print("\n" + "="*70)
    print("CATALOGO DE PRODUCTOS")
    print("="*70)
    
    # Recorremos todos los productos del diccionario e imprimimos cada profucto
    for producto in catalogo_productos.values():
        print(producto)
    print("="*70)


# Funcion para mostrar el historial completo de cambios
def ver_historial():

    # Verificamos si la lista de cambios esta vacia
    if not historial_cambios:
        print("No hay cambios registrados aun")
        return
    
    print("\n" + "="*70)
    print("HISTORIAL DE CAMBIOS")
    print("="*70)
    
    # Recorremos la lista de cambios con enumerate para obtener el indice
    for i, cambio in enumerate(historial_cambios, 1):
        print(f"{i}. {cambio}")
    
    print("="*70)


# Funcion para exportar el catalogo a un archivo Excel
def exportar_a_excel():

    # Verificamos si hay productos en el catalogo
    if not catalogo_productos:
        print("No hay productos para exportar")
        return
    
    # Definimos el nombre del archivo CSV temporal
    archivo_csv = 'catalogo_temp.csv'
    datos = [['Codigo', 'Nombre', 'Precio', 'Stock', 'Categoria', 'Fecha_Creacion']]
    
    # Recorremos todos los productos del catalogo y creamos una lista con estos datos
    for producto in catalogo_productos.values():
        fila = [
            producto.codigo,                                       
            producto.nombre,                                        
            producto.precio,                                        
            producto.stock,                                        
            producto.categoria,                                     
            producto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')  
        ]
        # Agregamos la fila a la lista de datos
        datos.append(fila)
    
    # Abrimos el archivo CSV
    with open(archivo_csv, 'w', newline='', encoding='utf-8') as f:
        escritor = csv.writer(f)
        escritor.writerows(datos)
    
    # Creamos un nuevo libro de excel, obtenemos una hoja activa y le asignamos el nombre/ 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Abrimos el archivo CSV que creamos anteriormente y con el lector de csv convertimos todo el contenido a una lista
    with open(archivo_csv, 'r', encoding='utf-8') as f:
        lector = csv.reader(f)
        datos_csv = list(lector)
    
    # aquí no le entendi muy bien al tutorial sobre como funciona, pero funciona
    for fila_num, fila_datos in enumerate(datos_csv, 1):
        for col_num, valor in enumerate(fila_datos, 1):
            if fila_num > 1 and col_num in [3, 4]:
                valor = float(valor) if col_num == 3 else int(valor)
            ws.cell(row=fila_num, column=col_num, value=valor)
    
    # Modificamos el formato con el cual se va a ver el excel
    font_encabezado = Font(bold=True, color="FFFFFF")  # Texto en negrita y blanco
    fill_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") #relleno de color azul para los cuadros
    alineacion = Alignment(horizontal="center", vertical="center")
    
    # Recorremos las primeras 6 columnas para aplicar los cambios
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = font_encabezado
        cell.fill = fill_encabezado
        cell.alignment = alineacion
    
    # Ajustamos automaticamente el ancho de cada columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Definimos el nombre del archivo Excel final y lo guardamos en la carpeta donde esta el codigo
    archivo_excel = 'catalogo_productos.xlsx'
    wb.save(archivo_excel)
    print(f"Catalogo exportado exitosamente a '{archivo_excel}'")
    
    # Creamos un mensaje para el historial
    cambio = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Se exporto el catalogo a Excel"
    historial_cambios.append(cambio)


# Funcion que muestra el menu principal con las opciones
def mostrar_menu():
 
    print("\n" + "="*50)
  
    print("SISTEMA DE INVENTARIO v2.2")
    print("="*50)
    print("1. Crear nuevo producto")
    print("2. Ver catalogo completo")
    print("3. Modificar stock de producto")
    print("4. Actualizar precio de producto")
    print("5. Ver historial de cambios")
    print("6. Exportar catalogo a Excel")  
    print("7. Salir")
    print("="*50)


# Funcion del menu para crear un nuevo producto
def menu_crear_producto():
  
    print("\n--- CREAR NUEVO PRODUCTO ---")
    
    # Solicitamos los datos del producto y eliminamos espacios en blanco o converrtimos los numeros a numeros decimales o enteros segun sea necesairio
    codigo = input("Codigo del producto: ").strip()
    nombre = input("Nombre: ").strip()
    precio = float(input("Precio: $"))
    stock = int(input("Stock inicial: "))
    categoria = input("Categoria: ").strip()
    
    # Llamamos a la funcion crear_producto con todos los datos ingresados
    crear_producto(codigo, nombre, precio, stock, categoria)


# Funcion del menu para modificar el stock
def menu_modificar_stock():

    print("\n--- MODIFICAR STOCK ---")
    
    codigo = input("Codigo del producto: ").strip()
    # Buscamos el producto en el catalogo y verificamos si se encuentra
    producto = buscar_producto(codigo)
    
    if not producto:
        print(f"No se encontro producto con codigo '{codigo}'")
        return
    
    print(f"Producto encontrado: {producto}")
    
    print("\n1. Agregar stock")
    print("2. Retirar stock")
    opcion = input("Selecciona operacion: ").strip()
    
    # Solicitamos la cantidad, la convertimos a entero y verificamos algunas cosas
    cantidad = int(input("Cantidad: "))
    
    if cantidad <= 0:
        # Si no lo es, mostramos error y terminamos
        print("La cantidad debe ser mayor a 0")
        return
    
    if opcion == "1":
        producto.modificar_stock(cantidad, 'agregar')
    elif opcion == "2":
        producto.modificar_stock(cantidad, 'retirar')
    else:
        print("Opcion no valida")


# Funcion del menu para actualizar precio
def menu_actualizar_precio():

    print("\n--- ACTUALIZAR PRECIO ---")

    codigo = input("Codigo del producto: ").strip()
    producto = buscar_producto(codigo)
    
    # Verificamos si se encontro el producto
    if not producto:
        print(f"No se encontro producto con codigo '{codigo}'")
        return
    
    print(f"Producto encontrado: {producto}")
    
    # Solicitamos el nuevo precio y lo convertimos a decimal
    nuevo_precio = float(input("Nuevo precio: $"))
    
    if nuevo_precio < 0:
        print("El precio no puede ser negativo")
        return
    
    producto.actualizar_precio(nuevo_precio)


# Funcion principal que ejecuta todo el programa
def main():
 
 
    print("Bienvenido al Sistema de Inventario!")
    
    # Iniciamos un ciclo infinito para el menu
    while True:
        # Mostramos el menu principal y solicitamos que el usuario elija una opcion
        mostrar_menu()
        
        opcion = input("\nSelecciona una opcion (1-7): ").strip()
        
        if opcion == "1":
            menu_crear_producto()
        elif opcion == "2":
            listar_productos()
        elif opcion == "3":
            menu_modificar_stock()
        elif opcion == "4":
            menu_actualizar_precio()
        elif opcion == "5":
            ver_historial()
        elif opcion == "6":
            exportar_a_excel()  
        elif opcion == "7":
            print("\nGracias por usar el sistema de inventario!")
            break
        # Si elige cualquier otra opcion
        else:
            print("Opcion no valida. Elige entre 1-7")
        
        input("\nPresiona Enter para continuar...")


if __name__ == "__main__":
    main()