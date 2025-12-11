import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# Clase que maneja la exportacion de datos a Excel
class ExportadorExcel:
    """
    Se encarga de realizar el manejo de los csv y creacion del archivo excel     
    """
    
    def exportar_catalogo(self, productos: list):
         # Verificamos si hay productos en el catalogo
        if not productos:
            print("No hay productos para exportar")
            return None
        
        # Esto corrige a la hora de nombrar los archivos
        try:
            # Definimos el nombre del archivo CSV temporal
            archivo_csv = 'catalogo_temp.csv'
            datos = [['Codigo', 'Nombre', 'Precio', 'Stock', 'Categoria', 'Fecha_Creacion']]
            
            # Recorremos todos los productos del catalogo y creamos una lista con estos datos
            for producto in productos:
                fila = [
                    producto._codigo,                                       
                    producto._nombre,                                        
                    producto._precio,                                        
                    producto._stock,                                        
                    producto._categoria,                                     
                    producto._fecha_creacion.strftime('%Y-%m-%d %H:%M:%S')  
                ]
                 # Agregamos la fila a la lista de datos
                datos.append(fila)

            # Abrimos el archivo CSV
            with open(archivo_csv, 'w', newline='', encoding='utf-8') as f:
                escritor = csv.writer(f)
                escritor.writerows(datos)

            # Creamos un nuevo libro de excel, obtenemos una hoja activa y le asignamos el nombre/ 
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            # Abrimos el archivo CSV que creamos anteriormente y con el lector de csv convertimos todo el contenido a una lista
            with open(archivo_csv, 'r', encoding='utf-8') as f:
                lector = csv.reader(f)
                datos_csv = list(lector)
            
            # aquÃ­ no le entendi muy bien al tutorial sobre como funciona, pero funciona
            for fila_num, fila_datos in enumerate(datos_csv, 1):
                for col_num, valor in enumerate(fila_datos, 1):
                    if fila_num > 1 and col_num in [3, 4]:
                        valor = float(valor) if col_num == 3 else int(valor)
                    ws.cell(row=fila_num, column=col_num, value=valor)
            
            # Aplica el formato a los encabezados
            self._aplicar_formato(ws)
            
            # Definimos el nombre del archivo Excel final y lo guardamos en la carpeta donde esta el codigo
            archivo_excel = 'catalogo_productos.xlsx'
            wb.save(archivo_excel)
            print(f"Catalogo exportado exitosamente a '{archivo_excel}'")
            
            return "Se exporto el catalogo a Excel"
        
        # soluciona errores al abrir el excel
        except PermissionError:
            print("Error: El archivo Excel esta abierto. Cierralo e intenta de nuevo")
            return None
        except Exception as e:
            print(f"Error al exportar a Excel: {e}")
            return None
    
    #  Metodo para aplicar formato al Excel
    def _aplicar_formato(self, ws):
        """ Aplica formato de colores y alineacion a los encabezados"""
         # Modifica el formato con el cual se va a ver el excel
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        alineacion = Alignment(horizontal="center", vertical="center")
        
         # Recorre las primeras 6 columnas para aplicar los cambios
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = font_encabezado
            cell.fill = fill_encabezado
            cell.alignment = alineacion

        # Ajusta automaticamente el ancho de cada columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width